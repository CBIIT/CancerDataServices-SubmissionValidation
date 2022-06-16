"""
validate that the xlsx file has the data and structure as expected
"""

import sys
import warnings
import pprint
with warnings.catch_warnings():
    warnings.simplefilter("ignore", category=UserWarning)
    from openpyxl import load_workbook
from cds_template_tool.models.errors import ValidateError  # noqa: E402
from cds_template_tool.helpers.logger import get_logger              # noqa: E402
from cds_template_tool.models.participant import Participant               # noqa: E402
from cds_template_tool.models.subfile import Subfile                         # noqa: E402
from cds_template_tool.models.study import Study                         # noqa: E402
from cds_template_tool.models.sample import Sample                   # noqa: E402
from cds_template_tool.models.errors import XlsxWorksheetError       # noqa: E402

logger = get_logger(__name__)


def validate_sheetnames(wb):
    """check sheet names to set expected in template, raise error if different set found"""

    found_names = wb.sheetnames
    expected_names = [
        "README",
        "Study",
        "Participant",
        "Sample",
        "File",
        "File-Participant-Sample Mapping",
        "Genomic Info",
        "Diagnosis (opt)",
        "Treatment (opt)",
        "Dictionary",
        "Terms and Value Sets"
    ]

    extra = list(set(found_names) - set(expected_names))
    missing = list(set(expected_names) - set(found_names))

    try:
        # if expected_names != found_names:
        if extra:
            extra_msg = "Extra sheets found: " + " ".join(extra)
            raise XlsxWorksheetError(extra_msg)
        if missing:
            missing_msg = "Missing sheets found: " + " ".join(missing)
            raise XlsxWorksheetError(missing_msg)
    except XlsxWorksheetError as err:
        logger.error("The xlsx file provided seems to have different sheets than the template")
        logger.error(err)
        sys.exit(200)



def check_sheetnames(wb):
    """ check sheet names """

    found_names = wb.sheetnames
    expected_names = [
        "Confirmed Variants",
        "Example Data",
        "Lists for Dropdowns",
        "Version",
    ]
    try:
        if expected_names != found_names:
            raise ValidateError
    except ValidateError:
        print("The xlsx file provided seems to have different sheets than the template")


def open_xlsx_file(xlsx_file):
    wb = load_workbook(filename=xlsx_file, data_only="True")
    return wb


def test_xlsx(xlsx_file):
    """ read and check/test """
    # read in
    _wb = open_xlsx_file(xlsx_file)  # noqa: F841
    pass


def check_worksheets(wb):
    """ check worksheets"""
    result_flag = None

    # check worksheets
    check_sheetnames(wb)

    # check variants
    check_variants(wb)

    # check dropdowns
    check_dropdown_lists(wb)

    return result_flag


def check_variants(wb):
    """check variants on 'confirmed variants' worksheet"""
    ws = wb["Confirmed Variants"]  # noqa: F841

    # check patient info

    # check snv info

    # check cnv info

    # check template info

    pass


def check_patient(object):
    pass


def check_cnv(object):
    pass


def check_dropdown_lists(wb):
    """check dropdowns on dropdown worksheet"""
    # check dropdown data
    ws = wb["Lists for Dropdowns"]

    found_list = [x.value for x in ws["A"]]
    print(found_list)

    pass

