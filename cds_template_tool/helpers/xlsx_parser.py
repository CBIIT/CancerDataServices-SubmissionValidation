"""
Read and parse submitted CDS Metadata Template xlsx files
uses openpyxl
"""
import sys
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
from openpyxl import load_workbook                             # noqa: E402
from cds_template_tool.helpers.logger import get_logger              # noqa: E402
from cds_template_tool.helpers.xlsx_validator import validate_sheetnames
from cds_template_tool.models.participant import Participant, max_col as max_par_col         # noqa: E402
from cds_template_tool.models.subfile import Subfile, max_col as max_subfile_col                      # noqa: E402
from cds_template_tool.models.f_p_s_mapping import F_p_s_mapping, max_col as max_fps_col         # noqa: E402
from cds_template_tool.models.genomic_info import Genomic_info, max_col as max_gi_col                      # noqa: E402
from cds_template_tool.models.study import Study                         # noqa: E402
from cds_template_tool.models.sample import Sample, max_col as max_samp_col                  # noqa: E402
from cds_template_tool.models.errors import XlsxWorksheetError       # noqa: E402

model_functions = {
    'File': [Subfile, max_subfile_col],
    'Sample': [Sample, max_samp_col],
    'Genomic Info': [Genomic_info, max_gi_col],
    'Participant': [Participant, max_par_col],
    'File-Participant-Sample Mapping':[F_p_s_mapping, max_fps_col]
}


logger = get_logger(__name__)

# TODO iterate over all known tabs in the excel file
# TODO mini class factory, function
# WARNING THIS IS A STUB - NOT FINISHED - BEING REFACTORED
def get_thingy(xlsx_file):
    holder = []
    # read in
    wb = load_workbook(filename=xlsx_file, data_only="True")

    # ensure workbook has all expected sheets - throws error if different
    validate_sheetnames(wb)

    subfiles_ = []

    # iterate of sheets
    for n, sheet in enumerate(wb.worksheets):
        logger.info('Sheet Index:[{}], Title:{}'.format(n, sheet.title))

        if sheet.title in model_functions:
            logger.debug('sheet name match')
            (cls, col) = model_functions[sheet.title]

            _rowdata = get_row_data(sheet, col)
            for row in _rowdata:
                # TODO capture in appropriate data structure
                subfiles_.append(cls(row))

    return subfiles_


def get_subfiles(xlsx_file):
    """ simple function to get the "Confirmed Variants" worksheet, which should have the data """
    subfiles = []

    # read in
    wb = load_workbook(filename=xlsx_file, data_only="True")

    # ensure workbook has all expected sheets - throws error if different
    validate_sheetnames(wb)

    subfile_ws = wb["File"]
    subfile_rowdata = get_row_data(subfile_ws, max_subfile_col)

    for row in subfile_rowdata:
        """just start with 1 subfile for now"""
        subfiles.append(Subfile(row))

    return subfiles

def get_samples(xlsx_file):
    """ simple function to get the "Confirmed Variants" worksheet, which should have the data """
    # read in
    wb = load_workbook(filename=xlsx_file, data_only="True")

    # ensure workbook has all expected sheets - throws error if different
    validate_sheetnames(wb)

    excel_file_tab_contents = wb["Sample"]
    sample_data = get_cell_data(excel_file_tab_contents,2,2,max_samp_col)

    """just start with 1 subfile for now"""
    my_sample = Sample(sample_data[0])
    return my_sample


def get_participants(xlsx_file):
    """ simple function to get the "Confirmed Variants" worksheet, which should have the data """
    # read in
    wb = load_workbook(filename=xlsx_file, data_only="True")

    # ensure workbook has all expected sheets - throws error if different
    validate_sheetnames(wb)

    excel_file_tab_contents = wb["Participant"]
    participant_data = get_cell_data(excel_file_tab_contents,2,2,max_par_col)

    """just start with 1 subfile for now"""
    my_participant = Participant(participant_data[0])

    return my_participant

def get_f_p_s_mapping(xlsx_file):
    """ simple function to get the "Confirmed Variants" worksheet, which should have the data """
    # read in
    wb = load_workbook(filename=xlsx_file, data_only="True")

    # ensure workbook has all expected sheets - throws error if different
    validate_sheetnames(wb)

    excel_file_tab_contents = wb["File-Participant-Sample Mapping"]
    fps_mapping_data = get_cell_data(excel_file_tab_contents,2,2,max_subfile_col)

    """just start with 1 subfile for now"""
    my_fps_mapping = F_p_s_mapping(fps_mapping_data[0])
    return my_fps_mapping


def get_genomic_infos(xlsx_file):
    """ simple function to get the "Confirmed Variants" worksheet, which should have the data """
    # read in
    wb = load_workbook(filename=xlsx_file, data_only="True")

    # ensure workbook has all expected sheets - throws error if different
    validate_sheetnames(wb)

    excel_file_tab_contents = wb["Genomic Info"]
    genomic_info_data = get_cell_data(excel_file_tab_contents,2,2,max_gi_col)

    """just start with 1 subfile for now"""
    my_gi = Genomic_info(genomic_info_data[0])
    return my_gi


def get_data_rows_between(openpyxl_worksheet, header_array, next_header_array=None):
    """
    look to find which rows have the headers in consecutive rows, in the first column
    and returns a tuple of what rows that should have data

    # does not check that rows are actually found (doesn't test for None)
    """

    starting_row = None
    ending_row = None

    # to make things easier to read
    ws = openpyxl_worksheet
    last_row = ws.max_row

    # starting row right after two headers
    for i in range(1, last_row):
        cell_id = "A" + str(i)
        next_cell_id = "A" + str(i + 1)

        if (ws[cell_id].value == header_array[0]
                and ws[next_cell_id].value == header_array[1]):  # noqa: W503
            # make sure next row exists ... (e.g. i+2)
            starting_row = i + 2
            break

    # find the last row of data section, by seeing when "next_header" starts....
    # if no "next_header" is specified, just use the last row in the spreadsheet, because
    # it is probably the last header on the sheet
    if next_header_array is None:
        ending_row = last_row
    # else look for when the next header starts and give the row right before it
    else:
        for j in range(starting_row, last_row):
            cell_id = "A" + str(j)
            if ws[cell_id].value == next_header_array[0]:
                ending_row = j - 1
                break
        # not checked here: what happens if next_header is not found

    # does not check if starting_row, or ending_row is still None (which isn't good...)
    return starting_row, ending_row


# todo: refactor above function to just return the header rows
#       then have another function to determine the data rows
#       can use the header rows to get and read headers
#       maybe have "section title" and "section headers" ?


def get_row_data(ws, end_column=0):
    """ read in actual data cells for each of the rows put in 2d list """
    """ max_col is the last known column 'R' """
    data = []
    for row_cells in ws.iter_rows( min_row=2, max_col=end_column, values_only=True):

        # this makes sure that you don't just have an entire row of Nones...
        if not all(x is None for x in row_cells):
            # data.extend(row_cells)
            data.append(row_cells)

    return data




def get_cell_data(ws, start_row, end_row,end_column=0):
    """ read in actual data cells for each of the rows put in 2d list """
    """ max_col is the last known column 'R' """
    data = []
    for row_cells in ws.iter_rows(
        min_row=start_row, max_col=end_column, max_row=end_row, values_only=True
    ):

        # this makes sure that you don't just have an entire row of Nones...
        if not all(x is None for x in row_cells):
            # data.extend(row_cells)
            data.append(row_cells)

    return data


def get_data(filename):
    """main entry point"""

    # open xlsx file and get an openpyxl worksheet that has the data
    ws = get_worksheet(filename)

    parsed_results_data = []
    parsed_results_types = []

    patient_header = ["Patient and Sample Information", "PSN"]
    snv_header = ["Confirmed SNV, Indels, MNV", "Gene Symbol"]
    cnv_header = ["Confirmed Copy Number Variants", "Gene Symbol"]
    fusion_header = ["Confirmed Fusions", "Variant ID"]

    patient_start, patient_end = get_data_rows_between(ws, patient_header, snv_header)
    patient_data = get_cell_data(ws, patient_start, patient_end)
    # TODO validation needs to throw error if patient_data is empty!
    my_patient = Patient(patient_data[0])  # just assume only one for now
    parsed_results_data.append(my_patient)
    parsed_results_types.append("patient")

    # SNV
    snv_start, snv_end = get_data_rows_between(ws, snv_header, cnv_header)
    all_snv_data = get_cell_data(ws, snv_start, snv_end)
    # now cast/put snv data into snv model:
    for a_snv_variant in all_snv_data:
        my_snv = Snv(a_snv_variant)
        # my_snv.printme()
        parsed_results_data.append(my_snv)
        parsed_results_types.append("snv")

    # CNV
    cnv_start, cnv_end = get_data_rows_between(ws, cnv_header, fusion_header)
    all_cnv_data = get_cell_data(ws, cnv_start, cnv_end)
    for a_cnv_variant in all_cnv_data:
        my_cnv = Cnv(a_cnv_variant)
        parsed_results_data.append(my_cnv)
        parsed_results_types.append("cnv")

    # Fusions
    fusion_start, fusion_end = get_data_rows_between(ws, fusion_header, None)
    all_fusion_data = get_cell_data(ws, fusion_start, fusion_end)
    for a_fusion_variant in all_fusion_data:
        my_fusion = Fusion(a_fusion_variant)
        parsed_results_data.append(my_fusion)
        parsed_results_types.append("fusion")

    # zip together
    # results = zip(parsed_results_types, parsed_results_data)
    # return results
    return parsed_results_types, parsed_results_data
